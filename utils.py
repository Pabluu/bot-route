from openpyxl import load_workbook
from address import address_remove

def path_file_XLSX(file):
  from pathlib import Path
  current = Path().resolve()

  return f'{current}/{file}'

def load_file(filename: str):
  # file = path_file_XLSX(filename)

  workbook = load_workbook(filename)
  worksheet = workbook.active

  return [workbook, worksheet]

def copy_sheet(ws) -> list:
  maxColumns = ws.max_column
  maxRows = ws.max_row
  copiedSheet = []

  for row in ws.iter_rows(max_row=maxRows,max_col=maxColumns, values_only=True):
    accumulated_items = []
    for item in row:
      accumulated_items.append(item)

    copiedSheet.append(accumulated_items)

  return copiedSheet

def index_column_address(data, address='Destination Address'):
  for index, item in enumerate(data[0]):
    if item == address:
      return index

def split_address(data, columnAddress, columnDescription):
  for indexRow, row in enumerate(data[1::], 1):
    addressData = data[indexRow][columnAddress]
    [address, description] = split_destination(addressData)

    data[indexRow][columnAddress] = address
    data[indexRow][columnDescription] = description

def split_destination(destination: str):
  from re import sub

  # removendo excesso de espaço
  destination = sub("\s\s+", " ", destination)

  [street, number, *description] = destination.strip().split(', ')

  if len(description) == 0:
    description = None
  else:
    description = description[0]

  address = f'{street}, {number}'

  return [address, description]

def insert_column(data, column, description, valueInitial=''):
  """
    Insere uma coluna nova na matriz e coloca '' no campo
  """

  for indexRow, row in enumerate(data, 0):
    data[indexRow].insert(column, valueInitial)

  data[0][column] = description

def organize_worksheet(data, columnDescription, columnAddress):
  """
    Função responsável por:
      1. Seperar o endereço em 2 partes:
        1: Rua e numero
        2: Referência
        
      2. Retirar o 'R ' no inicio do endereço substituindo por 'Rua '
      
  """

  # insere uma nova coluna
  insert_column(data, columnDescription, 'Address Line 2')

  # insere os endereços na nova coluna
  split_address(data, columnAddress, columnDescription)
  data[0][columnAddress] = 'Address Line 1'


def check_address(data: list, col: int):
  """Remove os endereços incorretos e insere os certos"""
  import re
  REGEX = r'^(r\. |r |rua )'

  for idxRow, row in enumerate(data[1::],1):
    adrs:str = row[col].lower()
    
    pattern = re.compile(REGEX, re.IGNORECASE)

    [address, number] = adrs.split(', ')
    
    if pattern.match(adrs):
      address = pattern.sub('Rua ', address)
      data[idxRow][col] = f'{address.title()}, {number}'

  return data

# def fetchAddress(data: list, col: int):
#   import requests
#   from os import getenv
#   from dotenv import load_dotenv
#   from urllib.parse import quote
  
#   load_dotenv()
#   GOOGLE = getenv('GOOGLE')
  
#   for idxRow, row in enumerate(data[1::],1):
#     street = quote(row[col])
#     print(row[col])
    
#     cityIndex = data[0].index('City')
#     city = row[cityIndex]

#     try:
#       URL = f'https://maps.googleapis.com/maps/api/geocode/json?address={street}&components=locality:{city}&key={GOOGLE}'
#       results = requests.get(URL).json()['results']

#       if results:
#         address_components = results[0].get('address_components', True)
#         isPolitical = 'politicial' in address_components[0].get('types', True)

#         if isPolitical:
#           continue
    
#       else:
#         autocomplete_url = f"https://maps.googleapis.com/maps/api/place/autocomplete/json?input={street}&location=-18.9439,-46.9929&radius=5000&key={GOOGLE}"
#         autocomplete_result = requests.get(autocomplete_url).json()

#         if autocomplete_result:
#           place_id = autocomplete_result['predictions'][0]['place_id']
#           details_url = f"https://maps.googleapis.com/maps/api/place/details/json?place_id={place_id}&key={GOOGLE}"

#         details_result = requests.get(details_url).json()['result']

#         location = details_result['geometry']['location']
#         # lat
#         data[idxRow][-2] = location['lat']

#         # lng
#         data[idxRow][-1] = location['lng']
    
#     except Exception as error:
#       print(error)
      
    
#     print(idxRow) # LOGGING

#   return data


def edit_length(data: list, col):
  """Remove o '1' na coluna de quantidade e o deixa sem nada """
  for line, row in enumerate(data[1::], 1):
    if row[col] == 1:
      data[line][col] = ''


def search_duplicated(data: list, address: str, columnAddress: str):
  """
    Retorna o index de um endereço em uma lista de outros endereços
  """
  for indexRow, row in enumerate(data[1::], 1):
    if(address in row[columnAddress]):
      return indexRow


def remove_duplicate(data, columnAddress, columnCount):
  """
    Aumenta a quantidade nos endereço repetido e remove os outros
  """

  newData = []
  listAddress = []
  newData.insert(0, data[0])

  idxRow = 1
  for indexRow, row in enumerate(data[1::], 1):
    address = row[columnAddress]

    countAddressOnList = listAddress.count(address)

    # Não existe na lista
    if(countAddressOnList == 0):
      newData.append(data[indexRow])
      listAddress.append(address)

      q = newData[idxRow][columnCount]
      idxRow += 1

    # Existe na lista(add +1)
    else:
      indxSD = search_duplicated(newData, address, columnAddress)
      q = newData[indxSD][columnCount]
      newData[indxSD][columnCount] = 1 + q

  return newData

def addAddress(data: list):
  # POR DO SOL
  
  por_do_sol = ['', '', '', '', 'Pôr Do Sol', '', 0, 'Congonha', 'Patrocínio', '38740000', '-18.92639', '-47.00644']
  data.append(por_do_sol)

  return data

def remove_apartment(data, columnAddress, columnLength):
  """
    Remove alguns endereços especificados no arquivos 'address'
  """

  indexRow = 1
  countAddress = 0

  for row in data[1::]:
    addressLowerCase: str = row[columnAddress].lower()
    address = addressLowerCase.split(', ')[0].replace('r ', '').replace('rua ', '')

    addressRemove = address_remove.get(address)

    if(addressRemove != None):
      del data[indexRow]
      countAddress += row[columnLength]

    else:
      indexRow += 1

  if(countAddress):
    data = addAddress(data)
    data[-1][6] = countAddress
    
  return data


def save_workbook(data, file):
  """salva as informações no arquivo"""

  from openpyxl import Workbook

  workbook = Workbook()
  worksheet = workbook.active

  for indexRow, row in enumerate(data):
    for indexColumn, column in enumerate(row):
      value = column

      worksheet.cell(row=indexRow+1, column=indexColumn+1, value=value)

  workbook.save(file)


def formatDateFile():
  '''
  Retorna o dia da semana com o sufixo
  '''

  from datetime import datetime
  from pytz import timezone
  import locale

  # Configura a localidade para português do Brasil
  locale.setlocale(locale.LC_TIME, 'pt_BR.utf8')
  tz = timezone('America/Sao_Paulo')

  day = datetime.now(tz).strftime("%A")

  if day in ['segunda', 'terça', 'quarta', 'quinta', 'sexta']:
    day += '-feira'

  return f'{day}.xlsx'